#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
txt_to_excel.py — แปลง text (paste / import ไฟล์) เป็น Excel อัตโนมัติ
Format ต่อบรรทัด: เครื่อง(2) ปี(4) เดือน(2) วัน(2) เวลา(4) EN(6) รวม 20 ตัวอักษร

ฟังก์ชันตรวจจับปัญหา 2 ชั้น:
  - ชั้นกฎตายตัว (ผิด = แดง) : อักขระแปลกปลอม / เดือน-วัน-เวลา-ปีผิดปกติ
  - ชั้นความถี่ต่อเครื่อง (สงสัย = เหลือง) : year/month/day ค่าที่ไม่ปกติ
    ของเครื่องนั้น (เครื่องนิ่งที่ค่าเดียว ~100%) — จับค่าตัวเลขล้วนแต่หลุด
    เช่น เดือน 18 / วัน 88 ที่กฎตายตัว/ความถี่ข้ามกัน  ส่วน time/en มี
    cardinality สูง (หลายค่า/unique ต่อรายการ) จึงไม่ใช้ความถี่ (กัน false
    positive ท่วม)

วิธีใช้:
  GUI : python txt_to_excel.py
  CLI : python txt_to_excel.py --file data.txt [--out result.xlsx]
"""
import calendar
import os
import re
import sys
import tkinter as tk
from collections import Counter, defaultdict
from tkinter import filedialog, messagebox, scrolledtext, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = ["เครื่อง", "ปี", "เดือน", "วัน", "เวลา", "EN"]
FIELD_LABELS = {"unit": "เครื่อง", "year": "ปี", "month": "เดือน",
                "day": "วัน", "time": "เวลา(TIME)", "en": "EN"}

# ฟิลด์ที่เครื่องมีค่า "นิ่ง" (~100%) ใช้ความถี่ต่อเครื่องได้
# time/en ไม่ใส่: cardinality สูง
FREQ_FIELDS = ["year", "month", "day"]

# อักขระที่ Excel เขียนไม่ได้ (control chars C0 + DEL + C1) — มักมาจาก OCR/สแกน
ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]")


def sanitize(text: str) -> str:
    """ลบ control characters ที่ Excel เขียนไม่ได้ออกจากข้อความ"""
    return ILLEGAL_CHARS_RE.sub("", text)


# --------------------------------------------------------------------------
# ฟังก์ชันตรวจสอบ
# --------------------------------------------------------------------------
def parse_line(line: str):
    """แยก 1 บรรทัดเป็น 6 ฟิลด์ + รายการปัญหากฎตายตัว
    คืน: (fields: dict, problems: list[str])"""
    line = sanitize(line)  # ลบ control chars ก่อนแยก (กัน openpyxl error)
    problems = []
    if len(line) != 20:
        problems.append(f"ความยาวไม่เท่ากับ 20 ตัว (ได้ {len(line)})")

    fields = {
        "unit":  line[0:2]  if len(line) >= 2  else "",
        "year":  line[2:6]  if len(line) >= 6  else "",
        "month": line[6:8]  if len(line) >= 8  else "",
        "day":   line[8:10] if len(line) >= 10 else "",
        "time":  line[10:14] if len(line) >= 14 else "",
        "en":    line[14:20] if len(line) >= 20 else "",
    }

    # 1) ฟิลด์ไหนมีอักขระไม่ใช่ตัวเลข
    for k, v in fields.items():
        if v and not v.isdigit():
            problems.append(f"{FIELD_LABELS[k]}มีอักขระแปลกปลอม ({v!r})")

    # 2) ปี ต้องเป็น 2000-2099
    y = fields["year"]
    if y.isdigit() and not (2000 <= int(y) <= 2099):
        problems.append(f"ปีผิดปกติ ({y})")

    # 3) เดือน 1-12
    m = fields["month"]
    if m.isdigit() and not (1 <= int(m) <= 12):
        problems.append(f"เดือนผิดปกติ ({m})")

    # 4) วัน ต้องตรงกับเดือนจริง (28-31 วัน)
    d = fields["day"]
    if d.isdigit():
        max_day = 31
        if m.isdigit() and 1 <= int(m) <= 12:
            yy = int(y) if y.isdigit() and 2000 <= int(y) <= 2099 else 2026
            max_day = calendar.monthrange(yy, int(m))[1]
        if not (1 <= int(d) <= max_day):
            problems.append(f"วันผิดปกติ ({d} — เดือนนี้มี {max_day} วัน)")

    # 5) เวลา(TIME) ต้องเป็น HHMM ที่มีจริง (00:00-23:59)
    t = fields["time"]
    if t.isdigit():
        if len(t) != 4:
            problems.append(f"เวลา(TIME)ความยาวผิด ({t})")
        else:
            hh, mm = int(t[:2]), int(t[2:])
            if not (0 <= hh <= 23 and 0 <= mm <= 59):
                problems.append(f"เวลา(TIME)ผิดปกติ ({t[:2]}:{t[2:]})")

    # 6) EN ต้องเป็นตัวเลข 6 หลัก
    en = fields["en"]
    if en.isdigit() and len(en) != 6:
        problems.append(f"EN ความยาวไม่เท่ากับ 6 ตัว ({en})")

    return fields, problems


def apply_frequency_suspicion(records):
    """ชั้นความถี่ต่อเครื่อง: ฟิลด์ year/month/day ที่ค่าของบรรทัดนี้
    ต่างจากค่าปกติของเครื่องนั้น (เครื่องนิ่งที่ค่าเดียว share>=0.5)
    → ขึ้น 'สงสัย' (ไม่ใช่ 'ผิด')
    records: list of (line_no, raw, fields, problems)
    คืน: list of (line_no, raw, fields, problems, suspicious: list[str])"""
    # 1) นับ distribution ต่อ (เครื่อง, ฟิลด์)
    norm = {}  # field -> unit -> Counter
    for _ln, _raw, fields, _problems in records:
        u = fields["unit"]
        for f in FREQ_FIELDS:
            v = fields.get(f)
            if not v:
                continue
            norm.setdefault(f, {}).setdefault(u, Counter())[v] += 1

    # 2) คำนวณค่าปกติ (dominant) + share ต่อเครื่อง-ฟิลด์
    #    คืน dict: field -> unit -> (dominant_value, share)
    dominant = {}
    for f, per_unit in norm.items():
        dominant[f] = {}
        for u, c in per_unit.items():
            total = sum(c.values())
            dom, domc = c.most_common(1)[0]
            dominant[f][u] = (dom, domc / total if total else 0)

    # 3) วน flag บรรทัดที่หลุดจากค่าปกติ
    enriched = []
    for ln, raw, fields, problems in records:
        u = fields["unit"]
        susp = []
        for f in FREQ_FIELDS:
            v = fields.get(f)
            if not v:
                continue
            entry = dominant[f].get(u)
            if not entry:
                continue
            dom, share = entry
            # เครื่องนิ่งที่ค่าเด่นชัด (share>=0.5) และค่านี้ต่างจากค่าเด่น
            if share >= 0.5 and v != dom:
                susp.append(f"{FIELD_LABELS[f]}ไม่ปกติของเครื่องนี้ (ค่าปกติคือ {dom})")
        enriched.append((ln, raw, fields, problems, susp))
    return enriched


def status_of(problems, suspicious):
    """จัดสถานะ 3 ระดับ: ผิด(แดง) > สงสัย(เหลือง) > ปกติ(เขียว)"""
    if problems:
        return "ผิด"
    if suspicious:
        return "สงสัย"
    return "ปกติ"


def problem_type(msg: str) -> str:
    """จัดกลุ่มประเภทปัญหา (ตัดค่าตัวเลขในวงเล็บทิ้ง)"""
    return re.sub(r"\s*\(.*?\)\s*$", "", msg).strip()


def detect_problem_machines(records):
    """ฟังก์ชันตรวจจับเครื่องที่มีปัญหา
    records: list of (line_no, raw, fields, problems, suspicious)
    คืน: dict เครื่อง -> {"total":int, "ok":int, "sus":int, "bad":int, "problems": Counter}"""
    machines = {}
    for _ln, _raw, fields, problems, suspicious in records:
        unit = fields["unit"] or "??"
        m = machines.setdefault(unit, {"total": 0, "ok": 0, "sus": 0, "bad": 0, "problems": Counter()})
        m["total"] += 1
        st = status_of(problems, suspicious)
        if st == "ผิด":
            m["bad"] += 1
            for p in problems:
                m["problems"][problem_type(p)] += 1
        elif st == "สงสัย":
            m["sus"] += 1
            for p in suspicious:
                m["problems"][problem_type(p)] += 1
        else:
            m["ok"] += 1
    return machines


# --------------------------------------------------------------------------
# สร้างไฟล์ Excel
# --------------------------------------------------------------------------
def build_excel(records, out_path, machines=None):
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    ok_fill = PatternFill("solid", fgColor="E2EFDA")      # เขียว
    sus_fill = PatternFill("solid", fgColor="FFF2CC")     # เหลือง
    bad_fill = PatternFill("solid", fgColor="FCE4EC")     # ชมพู

    # ---- ชีต 1: ข้อมูลทั้งหมด ----
    ws = wb.active
    ws.title = "ข้อมูลทั้งหมด"
    headers = HEADERS + ["สถานะ"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r, (line_no, raw, fields, problems, suspicious) in enumerate(records, 2):
        values = [fields["unit"], fields["year"], fields["month"],
                  fields["day"], fields["time"], fields["en"]]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.number_format = "@"
            cell.alignment = Alignment(horizontal="center")
        st = status_of(problems, suspicious)
        detail = problems if st == "ผิด" else suspicious
        status = "ปกติ" if st == "ปกติ" else f"{st}: " + " | ".join(detail)
        sc = ws.cell(row=r, column=7, value=status)
        sc.number_format = "@"
        sc.fill = ok_fill if st == "ปกติ" else (sus_fill if st == "สงสัย" else bad_fill)

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 30 if c == 7 else 12
    ws.freeze_panes = "A2"

    # ---- ชีต 2: เครื่องที่มีปัญหา ----
    if machines is None:
        machines = detect_problem_machines(records)
    ws2 = wb.create_sheet("เครื่องที่มีปัญหา")
    h2 = ["เครื่อง", "จำนวนรายการ", "ปกติ", "สงสัย", "มีปัญหา", "ประเภทปัญหา"]
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for unit in sorted(machines.keys()):
        m = machines[unit]
        if m["bad"] == 0 and m["sus"] == 0:
            continue
        detail = ", ".join(f"{t}x{n}" for t, n in m["problems"].most_common())
        vals = [unit, str(m["total"]), str(m["ok"]), str(m["sus"]), str(m["bad"]), detail]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=c, value=v)
            cell.number_format = "@"
            cell.alignment = Alignment(horizontal="center")
        # สีตามระดับรุนแรง
        fill = bad_fill if m["bad"] else sus_fill
        for c in range(1, 7):
            ws2.cell(row=row, column=c).fill = fill
        row += 1
    for c, w in zip(range(1, 7), (10, 12, 10, 10, 10, 55)):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.freeze_panes = "A2"

    # ---- ชีต 3: รายการผิดปกติ (ข้อมูลดิบ) ----
    ws3 = wb.create_sheet("รายการผิดปกติ")
    h3 = ["บรรทัดที่", "ข้อมูลดิบ", "เครื่อง", "สถานะ", "ปัญหา"]
    for c, h in enumerate(h3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for line_no, raw, fields, problems, suspicious in records:
        st = status_of(problems, suspicious)
        if st == "ปกติ":
            continue
        detail = problems if st == "ผิด" else suspicious
        vals = [str(line_no), raw, fields["unit"] or "??", st, " | ".join(detail)]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=c, value=v)
            cell.number_format = "@"
            cell.alignment = Alignment(horizontal="center" if c != 2 else "left")
        fill = bad_fill if st == "ผิด" else sus_fill
        for c in range(1, 6):
            ws3.cell(row=row, column=c).fill = fill
        row += 1
    for c, w in zip(range(1, 6), (10, 24, 10, 10, 75)):
        ws3.column_dimensions[get_column_letter(c)].width = w
    ws3.freeze_panes = "A2"

    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------
# ประมวลผลข้อความ (ใช้ร่วมกัน GUI / CLI)
# --------------------------------------------------------------------------
def process_text(text):
    """คืน (records, machines, summary_dict)"""
    records = []
    for i, raw in enumerate(text.splitlines(), 1):
        line = sanitize(raw.strip())
        if not line:
            continue
        fields, problems = parse_line(line)
        records.append((i, line, fields, problems))
    records = apply_frequency_suspicion(records)
    machines = detect_problem_machines(records)
    total = len(records)
    ok = sum(1 for *_r, problems, suspicious in records if status_of(problems, suspicious) == "ปกติ")
    sus = sum(1 for *_r, problems, suspicious in records if status_of(problems, suspicious) == "สงสัย")
    bad = total - ok - sus
    bad_machines = sorted(u for u, m in machines.items() if m["bad"] > 0 or m["sus"] > 0)
    summary = {"total": total, "ok": ok, "sus": sus, "bad": bad, "bad_machines": bad_machines}
    return records, machines, summary


# --------------------------------------------------------------------------
# GUI
# --------------------------------------------------------------------------
class App:
    def __init__(self, root):
        self.root = root
        root.title("txt → Excel อัตโนมัติ (ตรวจจับ: ผิด/สงสัย/ปกติ)")
        root.geometry("980x600")

        tk.Label(root, text="วางข้อมูล (หรือเปิดไฟล์ .txt) — format: เครื่อง(2) ปี(4) เดือน(2) วัน(2) เวลา(4) EN(6) รวม 20 ตัว",
                 anchor="w").pack(fill="x", padx=8, pady=(8, 2))

        # ---- พื้นที่หลัก: ซ้าย = กล่องข้อความ, ขวา = แผงสรุปเครื่อง ----
        main = tk.Frame(root)
        main.pack(fill="both", expand=True, padx=8, pady=4)

        self.txt = scrolledtext.ScrolledText(main, font=("Consolas", 10), wrap="none")
        self.txt.pack(side="left", fill="both", expand=True)

        # แผงสรุปเครื่อง (อัปเดตเรียลไทม์)
        panel = tk.Frame(main, width=300)
        panel.pack(side="right", fill="y", padx=(8, 0))
        panel.pack_propagate(False)

        tk.Label(panel, text="สรุปเครื่อง (อัปเดตอัตโนมัติ)", font=("Segoe UI", 10, "bold"),
                 anchor="w").pack(fill="x")

        self.tree = ttk.Treeview(panel, columns=("total", "ok", "sus", "bad"),
                                 show="tree headings", height=20)
        self.tree.heading("#0", text="เครื่อง")
        self.tree.heading("total", text="รวม")
        self.tree.heading("ok", text="ปกติ")
        self.tree.heading("sus", text="สงสัย")
        self.tree.heading("bad", text="ผิด")
        self.tree.column("#0", width=70, anchor="center")
        for col in ("total", "ok", "sus", "bad"):
            self.tree.column(col, width=45, anchor="center")
        self.tree.tag_configure("ok", background="#E2EFDA")
        self.tree.tag_configure("sus", background="#FFF2CC")
        self.tree.tag_configure("bad", background="#FCE4EC")
        tree_scroll = ttk.Scrollbar(panel, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="left", fill="y")

        self.summary_label = tk.Label(panel, text="", anchor="w", justify="left",
                                      fg="#1e8449", font=("Segoe UI", 9))
        self.summary_label.pack(fill="x", pady=(6, 0))

        # อัปเดตแผงสรุปสดทุกครั้งที่พิมพ์ / วาง / แก้ข้อความ
        self.txt.bind("<KeyRelease>", lambda e: self.refresh_overview())
        self.txt.bind("<<Paste>>", lambda e: self.refresh_overview())

        btns = tk.Frame(root)
        btns.pack(fill="x", padx=8, pady=4)
        tk.Button(btns, text="เปิดไฟล์ .txt", command=self.open_file).pack(side="left", padx=2)
        tk.Button(btns, text="สร้าง Excel", command=self.make_excel).pack(side="left", padx=2)
        tk.Button(btns, text="เคลียร์", command=self.clear_all).pack(side="left", padx=2)

        self.status = tk.Label(root, text="", anchor="w", justify="left", fg="#0b5394")
        self.status.pack(fill="x", padx=8, pady=(0, 8))

        self.refresh_overview()

    def get_text(self):
        return self.txt.get("1.0", "end")

    def refresh_overview(self):
        """ประมวลผลข้อความปัจจุบัน → อัปเดตแผงสรุปเครื่องแบบเรียลไทม์"""
        text = self.get_text().strip()
        if not text:
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.summary_label.config(text="ยังไม่มีข้อมูล")
            self.status.config(text="")
            return
        try:
            records, machines, summary = process_text(text)
        except Exception:
            return
        # ล้างตารางเดิม
        for item in self.tree.get_children():
            self.tree.delete(item)
        # ใส่ทีละเครื่อง เรียง: เครื่องมีปัญหาขึ้นก่อน แล้วเรียงตามเลขเครื่อง
        order = sorted(machines.keys(),
                       key=lambda u: (0 if (machines[u]["bad"] > 0 or machines[u]["sus"] > 0) else 1, u))
        for u in order:
            m = machines[u]
            tag = "bad" if m["bad"] else ("sus" if m["sus"] else "ok")
            self.tree.insert("", "end", text=u, values=(m["total"], m["ok"], m["sus"], m["bad"]), tags=(tag,))
        self.summary_label.config(
            text=(f"รวม {summary['total']} | ปกติ {summary['ok']} | "
                  f"สงสัย {summary['sus']} | ผิด {summary['bad']}"))
        bad = summary["bad_machines"]
        msg = f"พบ {summary['total']} รายการ | ปกติ {summary['ok']} | สงสัย {summary['sus']} | ผิด {summary['bad']}"
        if bad:
            msg += f" | เครื่องมีปัญหา: {', '.join(bad)}"
        self.status.config(text=msg, fg="#c0392b" if bad else "#1e8449")

    def open_file(self):
        path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
                data = f.read()
        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"เปิดไฟล์ไม่ได้: {e}")
            return
        self.txt.delete("1.0", "end")
        self.txt.insert("1.0", data)
        self.refresh_overview()

    def clear_all(self):
        self.txt.delete("1.0", "end")
        self.refresh_overview()

    def make_excel(self):
        text = self.get_text().strip()
        if not text:
            messagebox.showwarning("แจ้งเตือน", "กรุณาวางข้อมูลก่อน")
            return
        records, machines, summary = process_text(text)
        if not records:
            messagebox.showwarning("แจ้งเตือน", "ไม่พบข้อมูลที่ใช้ได้")
            return
        out = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="data_result.xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not out:
            return
        try:
            build_excel(records, out, machines)
        except Exception as e:
            messagebox.showerror("ผิดพลาด", f"สร้างไฟล์ไม่ได้: {e}")
            return
        bad = summary["bad_machines"]
        msg = f"สร้างไฟล์สำเร็จ: {out}\nรวม {summary['total']} รายการ (ปกติ {summary['ok']}, สงสัย {summary['sus']}, ผิด {summary['bad']})"
        if bad:
            msg += f"\n⚠ เครื่องที่มีปัญหา: {', '.join(bad)}"
        messagebox.showinfo("เสร็จสิ้น", msg)
        self.refresh_overview()


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------
def log(msg):
    """print ที่ปลอดภัยสำหรับ exe แบบ windowed (ไม่มี console)"""
    try:
        print(msg)
    except Exception:
        pass


def main_cli():
    args = sys.argv[1:]
    in_path = None
    out_path = None
    i = 0
    while i < len(args):
        if args[i] == "--file" and i + 1 < len(args):
            in_path = args[i + 1]; i += 2
        elif args[i] == "--out" and i + 1 < len(args):
            out_path = args[i + 1]; i += 2
        else:
            i += 1

    # exe แบบ windowed จะไม่มี console → ต้องโชว์ผลผ่าน popup แทน
    is_windowed_exe = getattr(sys, "frozen", False) and not getattr(sys.stdout, "isatty", lambda: True)()

    def notify(title, msg, is_error=False):
        log(msg)
        if is_windowed_exe:
            try:
                if is_error:
                    messagebox.showerror(title, msg)
                else:
                    messagebox.showinfo(title, msg)
            except Exception:
                pass

    if not in_path:
        notify("ผิดพลาด", "ใช้: txt_to_excel.exe --file data.txt [--out result.xlsx]", is_error=True)
        sys.exit(1)
    try:
        with open(in_path, "r", encoding="utf-8-sig", errors="replace") as f:
            text = f.read()
        records, machines, summary = process_text(text)
        if not out_path:
            base = os.path.splitext(in_path)[0]
            out_path = base + "_result.xlsx"
        build_excel(records, out_path, machines)
    except Exception as e:
        notify("ผิดพลาด", f"ไม่สามารถสร้างไฟล์ได้: {e}", is_error=True)
        sys.exit(1)
    msg = (f"สร้างไฟล์สำเร็จ: {out_path}\n"
           f"รวม {summary['total']} รายการ | ปกติ {summary['ok']} | สงสัย {summary['sus']} | ผิด {summary['bad']}")
    if summary["bad_machines"]:
        msg += "\nเครื่องที่มีปัญหา: " + ", ".join(summary["bad_machines"])
    notify("เสร็จสิ้น", msg)


if __name__ == "__main__":
    if "--file" in sys.argv:
        main_cli()
    else:
        root = tk.Tk()
        App(root)
        root.mainloop()
